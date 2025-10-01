import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

# ==============================
# 🔹 CONFIGURABLE VARIABLES
# ==============================

# Input & output file paths
input_csv = "report.csv"
output_xlsx = "report.xlsx"

# Static child headers (column names)
static_headers = [
    "From", "Project Name", "Branch",
    "Major", "Critical", "Minor", "Blocker",
    "Major", "Critical", "Minor", "Blocker",
    "Major", "Critical", "Minor", "Blocker",
    "Code Smells", "Vulnerabilities", "Bugs"
]

# Parent headers (spanning multiple child columns)
parent_headers = [
    "Categories", "", "", "",
    "Code Smells", "Code Smells", "Code Smells", "Code Smells",
    "Vulnerabilities", "Vulnerabilities", "Vulnerabilities", "Vulnerabilities",
    "Bugs", "Bugs", "Bugs", "Bugs",
    "Total Count", "Total Count"
]

# Column merging ranges for parent headers
merge_ranges = {
    "A1:A1": "Categories",
    "D1:G1": "Code Smells",
    "H1:K1": "Vulnerabilities",
    "L1:O1": "Bugs",
    "P1:R1": "Total Count"
}

# Styling configurations
parent_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue for parent headers
child_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for child headers
header_font = Font(bold=True, color="000000")  # Black text for both parent and child headers
header_alignment = Alignment(horizontal="center", vertical="center")

# Column width settings
child_header_width = 15  # Set uniform width for child headers

# Border settings
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# ==============================
# 🔹 PROCESS CSV DATA
# ==============================

# Read CSV file while keeping "N/A" values
df = pd.read_csv(input_csv, header=None, dtype=str)  # Preserve "N/A" values

# Assign dynamic column headers based on CSV content
df.columns = static_headers[:len(df.columns)]

# Save DataFrame to Excel with proper positioning
df.to_excel(output_xlsx, index=False, engine="openpyxl", startrow=1)  # Child headers on row 2

# ==============================
# 🔹 LOAD & FORMAT EXCEL
# ==============================

# Load workbook and select active sheet
wb = load_workbook(output_xlsx)
ws = wb.active

# ==============================
# 🔹 INSERT HEADERS & APPLY STYLES
# ==============================

# Insert Parent Headers (Row 1)
for col_num, header in enumerate(parent_headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Apply styles only to Parent Headers (Row 1)
for cell in ws[1]:
    cell.fill = parent_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border  # Apply border

# Apply styles only to Child Headers (Row 2)
for cell in ws[2]:
    cell.fill = child_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border  # Apply border

# Merge cells for parent headers
for merge_range, title in merge_ranges.items():
    ws.merge_cells(merge_range)
    ws[merge_range.split(":")[0]].value = title

# ==============================
# 🔹 APPLY BORDERS TO DATA CELLS
# ==============================

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border  # Apply border

# ==============================
# 🔹 ADJUST COLUMN WIDTH
# ==============================

# Set uniform width for all child header columns (Row 2)
for cell in ws[2]:  
    col_letter = cell.column_letter
    ws.column_dimensions[col_letter].width = child_header_width

# ==============================
# 🔹 CREATE TABLE & SAVE FILE
# ==============================

# Define table range dynamically based on data size
table_ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

# Create and style the table
table = Table(displayName="DataTable", ref=table_ref)
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False
)
table.tableStyleInfo = style
ws.add_table(table)

# Save the formatted Excel file
wb.save(output_xlsx)

print(f"Formatted Excel file created: {output_xlsx}")