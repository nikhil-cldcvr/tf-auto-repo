import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_csv("SonarQube_Metrics_Report.csv")

wb = Workbook()
ws = wb.active
ws.title = "SonarQube Report"

bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

row_index = 1
col_count = len(df.columns)

for i, row in df.iterrows():
    if "Report" in row[0]:  # Branch Report Header
        ws.append([row[0]])
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=col_count)
        ws[row_index][0].font = bold_font
        ws[row_index][0].alignment = Alignment(horizontal="center", vertical="center")
    else:
        ws.append(row.tolist())

        for col in range(1, col_count + 1):
            ws[row_index][col - 1].border = thin_border
            ws[row_index][col - 1].alignment = Alignment(horizontal="center", vertical="center")

    row_index += 1

# Adjust column width
for col in range(1, col_count + 1):
    ws.column_dimensions[get_column_letter(col)].width = 15

wb.save("SonarQube_Metrics_Report.xlsx")
